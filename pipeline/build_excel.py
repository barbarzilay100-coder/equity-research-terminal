#!/usr/bin/env python3
"""Builds docs/valuation-models.xlsx — the companion Excel workbook.

Three sheets:
  Assumptions — every DCF input as a labeled, named-range cell (blue = edit me)
  DCF         — 5-year FCF DCF for MSFT + WACC x terminal-growth sensitivity table
  Comps       — semiconductor trading comps repricing AVGO at peer medians

All model math lives in Excel formulas so the workbook recalculates when the
levers change; this script only injects snapshot inputs from data.json and
prints golden values (the same math in Python) for verification.

Run: python build_excel.py
"""
import json, os
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

BLUE, GREEN = "0000FF", "008000"
YELLOW = PatternFill("solid", fgColor="FFFF00")
USD_B, USD_PS, PCT, MULT = '$#,##0.0', '$#,##0.00', '0.0%', '0.0"x"'

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_XLSX = os.path.join(ROOT, "docs", "valuation-models.xlsx")

payload = json.load(open(os.path.join(ROOT, "data.json")))
co = {c["ticker"]: c for c in payload["companies"]}
msft, avgo = co["MSFT"], co["AVGO"]
peers = [c for c in payload["companies"]
         if c["sector"] == "Technology" and "Semiconductor" in c["industry"]
         and c["ticker"] != "AVGO" and (c.get("evEbitda") or 0) > 0 and (c.get("forwardPE") or 0) > 0]

wb = Workbook()

def put(ws, cell, value, fmt=None, color=None, bold=False, key=False, note=None):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Arial", size=11, bold=bold, color=color)
    if fmt: c.number_format = fmt
    if key: c.fill = YELLOW
    if note: c.comment = Comment(note, "build_excel.py")
    return c

# ---------------- Assumptions ----------------
ws = wb.active; ws.title = "Assumptions"
ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 22
put(ws, "A1", "Assumptions — Microsoft (MSFT) 5-year FCF DCF", bold=True)
put(ws, "A2", "Blue = hardcoded input (edit these) · Yellow = key levers · Black = formula", color=BLUE)
put(ws, "A4", "As of (data snapshot)"); put(ws, "B4", payload["generated"], color=BLUE)
put(ws, "A5", "Source"); put(ws, "B5", "Yahoo Finance via yfinance pipeline (data.json)", color=BLUE)
put(ws, "A7", "Base FCF ($B, trailing)"); put(ws, "B7", msft["fcf"], USD_B, color=BLUE,
    note=f"freeCashflow, snapshot {payload['generated']}. Trailing FCF is capex-depressed: net income is ${msft['netIncome']}B.")
put(ws, "A8", "FCF growth, years 1-5"); put(ws, "B8", 0.10, PCT, color=BLUE, key=True,
    note=f"Deliberate lever. Snapshot trailing revenue growth is {msft['revGrowth']}% — 10% assumes FCF grows below revenue while capex stays elevated.")
put(ws, "A9", "Terminal growth"); put(ws, "B9", 0.025, PCT, color=BLUE, key=True,
    note="Long-run nominal GDP-ish growth; must stay below WACC.")
put(ws, "A10", "WACC"); put(ws, "B10", 0.09, PCT, color=BLUE, key=True,
    note="Mega-cap equity discount rate assumption; flex in the sensitivity table.")
put(ws, "A11", "Net debt ($B) = debt − cash"); put(ws, "B11", round(msft["debt"] - msft["cash"], 3), USD_B, color=BLUE,
    note=f"totalDebt {msft['debt']}B − totalCash {msft['cash']}B, snapshot.")
put(ws, "A12", "Market cap ($B)"); put(ws, "B12", msft["marketCap"], USD_B, color=BLUE)
put(ws, "A13", "Share price ($)"); put(ws, "B13", msft["price"], USD_PS, color=BLUE)
put(ws, "A14", "Shares out (B)"); put(ws, "B14", "=MktCap/MktPrice", '#,##0.000')
for name, ref in [("BaseFCF", "$B$7"), ("GrowthY15", "$B$8"), ("TermGrowth", "$B$9"), ("WACC", "$B$10"),
                  ("NetDebt", "$B$11"), ("MktCap", "$B$12"), ("MktPrice", "$B$13"), ("SharesOut", "$B$14")]:
    wb.defined_names.add(DefinedName(name, attr_text=f"Assumptions!{ref}"))

# ---------------- DCF ----------------
ws = wb.create_sheet("DCF")
ws.column_dimensions["A"].width = 36
for col in "BCDEF": ws.column_dimensions[col].width = 13
put(ws, "A1", "DCF — Microsoft (MSFT), 5-year free-cash-flow model", bold=True)
put(ws, "A3", "", bold=True)
for t in range(1, 6):
    put(ws, f"{get_column_letter(1+t)}3", f"Year {t}", bold=True)
put(ws, "A4", "FCF ($B)")
put(ws, "A5", "Discount factor")
put(ws, "A6", "PV of FCF ($B)")
for t in range(1, 6):
    col = get_column_letter(1 + t)
    put(ws, f"{col}4", f"=BaseFCF*(1+GrowthY15)^{t}", USD_B, color=GREEN)
    put(ws, f"{col}5", f"=1/(1+WACC)^{t}", '0.000', color=GREEN)
    put(ws, f"{col}6", f"={col}4*{col}5", USD_B)
put(ws, "A8", "Sum PV, years 1-5 ($B)");            put(ws, "B8", "=SUM(B6:F6)", USD_B)
put(ws, "A9", "Terminal value at year 5 ($B)");     put(ws, "B9", "=F4*(1+TermGrowth)/(WACC-TermGrowth)", USD_B, color=GREEN)
put(ws, "A10", "PV of terminal value ($B)");        put(ws, "B10", "=B9*F5", USD_B)
put(ws, "A11", "Enterprise value ($B)");            put(ws, "B11", "=B8+B10", USD_B)
put(ws, "A12", "Less: net debt ($B)");              put(ws, "B12", "=NetDebt", USD_B, color=GREEN)
put(ws, "A13", "Equity value ($B)");                put(ws, "B13", "=B11-B12", USD_B)
put(ws, "A14", "Implied value per share ($)");      put(ws, "B14", "=B13/SharesOut", USD_PS, bold=True, key=True)
put(ws, "A15", "Market price ($)");                 put(ws, "B15", "=MktPrice", USD_PS, color=GREEN)
put(ws, "A16", "Implied upside/(downside)");        put(ws, "B16", "=B14/B15-1", PCT, bold=True)
put(ws, "A18", "Reading the gap: trailing FCF is capex-depressed (net income ~3x FCF in the snapshot), so a", color=BLUE)
put(ws, "A19", "conservative trailing-FCF DCF reads far below market — the price embeds AI-cycle FCF normalization.", color=BLUE)
put(ws, "A20", "Flex the growth lever (Assumptions!B8) and the table below to see what is being priced in.", color=BLUE)

put(ws, "A22", "Sensitivity — implied value per share ($): WACC (rows) × terminal growth (columns)", bold=True)
gs = [0.015, 0.02, 0.025, 0.03, 0.035]
ws_rows = [0.08, 0.085, 0.09, 0.095, 0.10]
for j, g in enumerate(gs):
    put(ws, f"{get_column_letter(2+j)}23", g, PCT, bold=True)
for i, w in enumerate(ws_rows):
    r = 24 + i
    put(ws, f"A{r}", w, PCT, bold=True)
    for j in range(len(gs)):
        col = get_column_letter(2 + j)
        terms = "+".join(f"BaseFCF*(1+GrowthY15)^{t}/(1+$A{r})^{t}" for t in range(1, 6))
        tv = f"BaseFCF*(1+GrowthY15)^5*(1+{col}$23)/(($A{r}-{col}$23)*(1+$A{r})^5)"
        put(ws, f"{col}{r}", f"=({terms}+{tv}-NetDebt)/SharesOut", USD_PS)

# ---------------- Comps ----------------
ws = wb.create_sheet("Comps")
ws.column_dimensions["A"].width = 9; ws.column_dimensions["B"].width = 32
for col in "CDEFGHI": ws.column_dimensions[col].width = 13
put(ws, "A1", "Trading comps — Broadcom (AVGO) vs. semiconductor peers", bold=True)
put(ws, "A2", f"Snapshot {payload['generated']} from the pipeline (data.json). Peers = universe members with industry "
              f"'Semiconductors' — a finer peer set than the sector-level medians the site computes.", color=BLUE)
headers = ["Ticker", "Company", "Price ($)", "Mkt Cap ($B)", "EV ($B)", "EV/EBITDA (x)", "Fwd P/E (x)", "EBITDA ($B)", "Fwd EPS ($)"]
for j, h in enumerate(headers):
    put(ws, f"{get_column_letter(1+j)}4", h, bold=True)
r = 5
for p in peers:
    put(ws, f"A{r}", p["ticker"]); put(ws, f"B{r}", p["name"])
    put(ws, f"C{r}", p["price"], USD_PS, color=BLUE); put(ws, f"D{r}", p["marketCap"], USD_B, color=BLUE)
    put(ws, f"E{r}", p["ev"], USD_B, color=BLUE); put(ws, f"F{r}", p["evEbitda"], MULT, color=BLUE)
    put(ws, f"G{r}", p["forwardPE"], MULT, color=BLUE)
    put(ws, f"H{r}", f"=E{r}/F{r}", USD_B); put(ws, f"I{r}", f"=C{r}/G{r}", USD_PS)
    r += 1
last = r - 1
put(ws, f"A{r}", "Median", bold=True); put(ws, f"B{r}", f"Peer median ({len(peers)} names, excludes AVGO)", bold=True)
put(ws, f"F{r}", f"=MEDIAN(F5:F{last})", MULT, bold=True); put(ws, f"G{r}", f"=MEDIAN(G5:G{last})", MULT, bold=True)
med_row = r
r += 2
put(ws, f"A{r}", "AVGO"); put(ws, f"B{r}", avgo["name"], bold=True)
put(ws, f"C{r}", avgo["price"], USD_PS, color=BLUE); put(ws, f"D{r}", avgo["marketCap"], USD_B, color=BLUE)
put(ws, f"E{r}", avgo["ev"], USD_B, color=BLUE); put(ws, f"F{r}", avgo["evEbitda"], MULT, color=BLUE)
put(ws, f"G{r}", avgo["forwardPE"], MULT, color=BLUE)
put(ws, f"H{r}", f"=E{r}/F{r}", USD_B); put(ws, f"I{r}", f"=C{r}/G{r}", USD_PS)
avgo_row = r
r += 2
blk = [
    ("Net debt ($B) = debt − cash", round(avgo["debt"] - avgo["cash"], 3), USD_B, BLUE,
     f"totalDebt {avgo['debt']}B − totalCash {avgo['cash']}B, snapshot."),
    ("Shares out (B)", f"=D{avgo_row}/C{avgo_row}", '#,##0.000', None, None),
    ("Implied EV @ median EV/EBITDA ($B)", f"=F{med_row}*H{avgo_row}", USD_B, None, None),
    ("Implied equity ($B)", f"=B{r+2}-B{r}", USD_B, None, None),
    ("Implied price — EV/EBITDA route ($)", f"=B{r+3}/B{r+1}", USD_PS, None, None),
    ("Implied price — Fwd P/E route ($)", f"=G{med_row}*I{avgo_row}", USD_PS, None, None),
    ("Blended implied value ($)", f"=AVERAGE(B{r+4}:B{r+5})", USD_PS, None, None),
    ("Market price ($)", f"=C{avgo_row}", USD_PS, None, None),
    ("Implied upside/(downside)", f"=B{r+6}/B{r+7}-1", PCT, None, None),
]
put(ws, f"A{r-1}", "Implied valuation of AVGO at peer medians", bold=True)
for k, (label, val, fmt, color, note) in enumerate(blk):
    put(ws, f"A{r+k}", label)
    put(ws, f"B{r+k}", val, fmt, color=color, bold=(k in (6, 8)), key=(k == 6), note=note)

wb.save(OUT_XLSX)

# ---------------- Golden values (same math in Python) ----------------
def dcf_ps(wacc, g5, gt):
    pv = sum(msft["fcf"] * (1 + g5) ** t / (1 + wacc) ** t for t in range(1, 6))
    tv = msft["fcf"] * (1 + g5) ** 5 * (1 + gt) / (wacc - gt)
    ev = pv + tv / (1 + wacc) ** 5
    shares = msft["marketCap"] / msft["price"]
    return (ev - (msft["debt"] - msft["cash"])) / shares

med = lambda xs: sorted(xs)[len(xs) // 2] if len(xs) % 2 else sum(sorted(xs)[len(xs) // 2 - 1:len(xs) // 2 + 1]) / 2
m_ee, m_fp = med([p["evEbitda"] for p in peers]), med([p["forwardPE"] for p in peers])
ebitda = avgo["ev"] / avgo["evEbitda"]
shares = avgo["marketCap"] / avgo["price"]
ip_ee = (m_ee * ebitda - (avgo["debt"] - avgo["cash"])) / shares
ip_fp = m_fp * avgo["price"] / avgo["forwardPE"]
print(f"GOLDEN  DCF/share base case      : {dcf_ps(0.09, 0.10, 0.025):9.2f}")
print(f"GOLDEN  DCF/share (8%,3.5%)      : {dcf_ps(0.08, 0.10, 0.035):9.2f}")
print(f"GOLDEN  DCF/share (10%,1.5%)     : {dcf_ps(0.10, 0.10, 0.015):9.2f}")
print(f"GOLDEN  comps median EV/EBITDA   : {m_ee:9.3f}")
print(f"GOLDEN  comps median Fwd P/E     : {m_fp:9.3f}")
print(f"GOLDEN  AVGO implied EV/EBITDA   : {ip_ee:9.2f}")
print(f"GOLDEN  AVGO implied Fwd P/E     : {ip_fp:9.2f}")
print(f"GOLDEN  AVGO blended / upside    : {(ip_ee+ip_fp)/2:9.2f} / {((ip_ee+ip_fp)/2/avgo['price']-1)*100:6.1f}%")
print(f"peers: {len(peers)}  wrote docs/valuation-models.xlsx")
